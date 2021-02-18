# 셀 이동
from openpyxl import load_workbook
wb = load_workbook("sample2.xlsx")
ws = wb.active

# 번호 영어 수학 -> 번호 국어 영어 수학
# B1부터 C11까지를 한열 오른쪽으로 이동
ws.move_range("B1:C11", rows=0, cols=1)
ws["B1"].value = "국어"       # B1 셀에 국어 입력

ws.move_range("D1:D11", rows=2, cols=-2)

wb.save("sample_move.xlsx")