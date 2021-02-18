# 파일 열기
from openpyxl import load_workbook      # 파일 불러오기
wb = load_workbook("sample.xlsx")
ws = wb.active      # 활성화된 Sheet

# cell의 개수를 모를 때
for x in range(1, ws.max_row + 1):           # 워크시트가 가진 최대 row 수
    for y in range(1, ws.max_column + 1):    # 워크시트가 가진 최대 col 수
        print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 ..
    print()