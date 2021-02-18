# 셀 영역
from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):                                      # 10 줄 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]     # 영어 column 만 가지고 오기
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"]       # 영어, 수학 column 함께 가지고 오기
for cols in col_range:      # 한개의 column 씩 가져온다.
    for cell in cols:
        print(cell.value, end=" ")
    print()

row_title = ws["1"]     # 첫 번째 row만 가지고 온다
for cell in row_title:
    print(cell.value, end=" ")
print()

row_range = ws[2:6]   # slicing과 다르게 2부터 6까지 포함해서 가져온다.
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

print()


# cell의 위치에 대한 정보 출력
from openpyxl.utils.cell import coordinate_from_string  # coordinate_from_string 함수를 사용

row_range = ws[2:ws.max_row]    # 2번째 행부터 마지막 행까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        print(cell.coordinate, end=" ")     # A2, B2, C2 등등 각 셀의 위치를 나타낸다.
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")     # ('A', 2)  ('B', 2) 등 튜플로 나타낸다.
    print()


# 전체 rows
print(tuple(ws.rows))   # (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>), ........ 각 행의 열원소를 튜플 하나씩으로 묶었다.

for row in tuple(ws.rows):      # (A1,B1,C1), (A2,B2,C2), (A3,B3,C3)
    print(row[1].value, end=" ")    # B1, B2, B3, B4, B5..... 의 값 출력
print()

# 전체 columns
print(tuple(ws.columns))    # 각 열을 튜플 하나씩으로 묶었다.

for col in tuple(ws.columns):      # (A1,A2,A3...), (B1,B2,B3....), (C1,C2,C3...)
    print(col[0].value, end=" ")    # A1, B1, C1 의 값 출력
print()


# iter를 사용해서 범위를 제한해서 가져오는 방법
# min은 지정하지 않으면 가장 작은 값, max는 지정하지 않으면 가장 큰 값

# 1~5행, 2~3열(B,C)
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value, row[1].value)    # B1~B5, C1~C5까지 가져온다. row[0]은 B이기 때문이다.

for col in ws.iter_cols(min_row=1, max_row=5, min_col=2, max_col=3):
    print(col)  # (<Cell 'Sheet'.B1>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.B3>, <Cell 'Sheet'.B4>, <Cell 'Sheet'.B5>)

wb.save("sample2.xlsx")
