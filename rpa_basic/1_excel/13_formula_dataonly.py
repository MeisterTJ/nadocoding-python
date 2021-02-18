# 수식 (데이터전용)
from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx")
ws = wb.active

# 수식이 있는 cell의 경우 데이터가 아닌 수식을 출력하는 문제가 있다.
for row in ws.values:
    for cell in row:
        print(cell)

# 수식이 아닌 실제 데이터만 가지고 온다.
# openpyxl을 통해서 만들어진 엑셀에 수식이 들어갔다면 사람이 한번 열어서 다시 저장하는게 좋다. 
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active
for row in ws.values:
    for cell in row:
        print(cell)     # 수식이 들어갔지만, 계산이 되지 않는 상태의 데이터는 None이라고 표시된다.